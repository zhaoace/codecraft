#!/usr/bin/env python
# coding=utf-8

import os

from openpyxl import load_workbook
from openpyxl import Workbook


def example_reading_excel():
    wb = load_workbook(filename='balances.xlsx', read_only=True)
    ws = wb['MySheet']

    for row in ws.rows:
        for cell in row:
            print(cell.value)

def example_writing_excel():
    wb = Workbook()
    ws = wb.create_sheet("MySheet")
    d = ws.cell(row=4, column=2, value=10)
    wb.save('balances.xlsx')


def test_writing_and_reading():
    example_writing_excel()
    example_reading_excel()
    if os.path.isfile('balances.xlsx') :
        print "remove temp file 'balances.xlsx'"
        os.remove('balances.xlsx')



if __name__ == '__main__':
    test_writing_and_reading()
