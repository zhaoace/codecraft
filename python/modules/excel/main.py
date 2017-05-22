#!/usr/bin/env python
# coding=utf-8

import os

from openpyxl import load_workbook
from openpyxl import Workbook

TEMP_EXCEL_NAME='example.xlsx'


def example_reading_excel():
    wb = load_workbook(filename=TEMP_EXCEL_NAME, read_only=True)
    ws = wb.active

    for row in ws.rows:
        for cell in row:
            print(cell.value)

def example_writing_excel():
    wb = Workbook()
    ws = wb.active
    for i in range(1,5):
        d = ws.cell(row=i, column=1, value=10)
    wb.save(TEMP_EXCEL_NAME)


def test_writing_and_reading():
    example_writing_excel()
    example_reading_excel()
    if os.path.isfile(TEMP_EXCEL_NAME) :
        print "remove temp file TEMP_EXCEL_NAME"
        os.remove(TEMP_EXCEL_NAME)



if __name__ == '__main__':
    test_writing_and_reading()
